import tkinter as tk
from tkinter import ttk
import tkinter.font as tkFont
from tkinter import messagebox
import openpyxl
import pandas as pd

class fundOperationNumber():

    def __init__(self, master):
        self.master = master
        self.frame = tk.Frame(master)
        self.frame.pack(fill="x") # fill all the parent's size (row)
        self.fundNumberView = tk.StringVar()
        self.numberValue = tk.StringVar()
        self.fundYear = tk.StringVar()
        self.fundMonth = tk.StringVar()
        self.fundDay = tk.StringVar()
        self.year = tk.StringVar()
        self.month = tk.StringVar()
        self.day = tk.StringVar()
        self.fundOperationFileInstance = []

    def fundMsgYear(self,*args):
        # 年份获取
        #print(self.yearChosen.get())
        self.fundYear.set(self.yearChosen.get())
        yearIndex = (int)(len(self.fundOperationFileInstance))
        for timeIndex in range(yearIndex):
            # 遍历，统一修改各写入基金的定投年份
            self.fundOperationFileInstance[timeIndex].yearChosen.set(self.fundYear.get())
            #self.fundOperationFileInstance[timeIndex].fundYear.set(self.fundYear.get())



    def fundMsgMonth(self,*args):
        # 月份获取
        print(self.monthChosen.get())
        self.fundMonth.set(self.monthChosen.get())
        monthIndex = (int)(len(self.fundOperationFileInstance))
        for timeIndex in range(monthIndex):
            self.fundOperationFileInstance[timeIndex].monthChosen.set(self.fundMonth.get())

    def fundMsgDay(self,*args):
        # 天数获取
        print(self.dayChosen.get())
        self.fundDay.set(self.dayChosen.get())
        dayIndex = (int)(len(self.fundOperationFileInstance))
        for timeIndex in range(dayIndex):
            self.fundOperationFileInstance[timeIndex].dayChosen.set(self.fundDay.get())

    def fundMsgNumberView(self, *args):
        # 显示几类基金框输入
        #print(self.numberViewChosen.get())
        #self.fundNumberView.set(self.numberViewChosen.get())
        self.fundNumberView.set(self.numberViewChosen.get())
        self.addFundNum()

    def fundNumberViewSet(self):
        # software introduction
        customFont = tkFont.Font(family="Arial Narrow",size=15, weight=tkFont.BOLD,slant=tkFont.ITALIC)
        self.recordValue = tk.Label(self.frame,font=customFont,width=60, fg="green", text="This software is made by Galen. ")
        self.recordValue.grid(row=0, column=1)
        # 选择基金输入记录数
        self.recordValue = tk.Label(self.frame,width=20 ,text="选择基金输入记录数:")
        self.recordValue.grid(row=1, column=0)

        self.numberViewChosen = tk.ttk.Combobox(self.frame, width=3, textvariable=self.numberValue, state='readonly')
        self.numberViewChosen['values'] = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10') # 设置下拉列表的值
        self.numberViewChosen.place(x=135, y = 30)  # 设置其在界面中出现的位置 column代表列 row 代表行
        self.numberViewChosen.current(0)
        self.numberViewChosen.bind("<<ComboboxSelected>>",self.fundMsgNumberView)
        self.fundOperationFileInstance.append(fundOperationFile(self.master))
        self.fundOperationFileInstance[0].fundoperation(5)


        # 定投时间
        self.time = tk.Label(self.frame,justify="left", text="统一设置定投时间: ")
        self.time.place(x=300, y=30) # execute to locate

        # 定投时间设置
        endYear = 3001 # not include 3001
        startYear = 2018
        n = endYear - startYear
        yearList=[None] * n
        yearIndex = 0
        for i in range(startYear,endYear):
            yearList[yearIndex] = i
            if i < (endYear -1):  # end year that we don't need to add 1
                yearIndex = yearIndex + 1;

        self.yearChosen = tk.ttk.Combobox(self.frame, width=5, textvariable=self.year, state='readonly')
        self.yearChosen['values'] = yearList # 设置下拉列表的值
        self.yearChosen.place(x=415, y=30) # 设置其在界面中出现的位置 column代表列 row 代表行
        self.yearChosen.current(1) # 设置下拉列表默认显示的值，1为yearChosen['values'] 的下标值
        self.yearChosen.bind("<<ComboboxSelected>>",self.fundMsgYear)  #绑定事件,(下拉列表框被选中时，绑定fundMsgYear()函数)

        self.monthChosen = tk.ttk.Combobox(self.frame, width=3, textvariable=self.month, state='readonly')
        self.monthChosen['values'] = ('1','2','3','4','5','6','7','8','9','10','11','12') # 设置下拉列表的值
        self.monthChosen.place(x=470, y=30) # 设置其在界面中出现的位置 column代表列 row 代表行
        self.monthChosen.current(0)
        self.monthChosen.bind("<<ComboboxSelected>>",self.fundMsgMonth)

        self.dayChosen = tk.ttk.Combobox(self.frame, width=3, textvariable=self.day, state='readonly')
        self.dayChosen['values'] = ('1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31') # 设置下拉列表的值
        self.dayChosen.place(x=513, y=30) # 设置其在界面中出现的位置 column代表列 row 代表行
        self.dayChosen.current(0)
        self.dayChosen.bind("<<ComboboxSelected>>",self.fundMsgDay)

        # 写入基金数据到Excel
        self.writeExcel = tk.Button(self.frame, text="提交基金定投记录到Excel", command=self.fundWriteExcel)
        self.writeExcel.place(x=650, y=30)

        # 换行
        self.seperator = tk.Label(self.frame,justify="left", text="  ")
        self.seperator.grid(row=2, column=0)

    def fundWriteExcel(self):
        # 提交定投基金记录
        dayIndex = (int)(len(self.fundOperationFileInstance))
        self.commitData = []
        for timeIndex in range(dayIndex):
            self.tmpData = []
            self.data =self.fundOperationFileInstance[timeIndex]
            # 定投时间
            self.tmpData.append(self.data.yearChosen.get() + "." + self.data.monthChosen.get() + '.' + self.data.dayChosen.get())
            #基金代码
            self.tmpData.append(self.data.fundCodeChosen.get())
            #基金名称详情
            self.tmpData.append(self.data.fundCodeDetailTxt.get())
            # 买入卖出金额
            self.tmpData.append(self.data.moneySet.get())
            # 成交单价
            self.tmpData.append(self.data.onePriceSet.get())
            # 盈利收益率
            self.tmpData.append(self.data.earningRateSet.get())
            # 市盈率
            self.tmpData.append(self.data.peRatioSet.get())
            self.commitData.append(self.tmpData)
        print(list(self.commitData)) # 打印定投基金记录

        # 定投基金记录写入到Excel表中
        self.excelData = openpyxl.load_workbook(r'C:\Users\Galen\Desktop\python_shell\定投记录.xlsx')
        # print(data.get_named_ranges()) # 输出工作页索引范围
        # print(data.get_sheet_names()) # 输出所有工作页的名称
        # 取第一张表
        sheetname = self.excelData.sheetnames  # wb.sheetnames
        table = self.excelData[sheetname[0]]  # wb[sheetname]
        #print(table.title)  # 输出表名
        nrows = table.max_row  # 获得行数
        column = 2
        for valueIndex in self.commitData:
            print("valueIndex = ",valueIndex)
            for value in valueIndex:
                table.cell(nrows + 1, column).value = value
                column = column + 1
            # 重设开始定位位置
            nrows = nrows + 1
            column = 2
        self.excelData.save(r'C:\Users\Galen\Desktop\python_shell\定投记录.xlsx')
        tk.messagebox.showinfo('提示', '写入成功')  #  状态提示

    # set the number of fund to add to my app
    def addFundNum(self):
        numbers =  (int)(self.fundNumberView.get())

        while self.fundOperationFileInstance:
            deleteWidget = self.fundOperationFileInstance.pop()
            '''
            # 遍历控件，delete class's frame 的 控件
            for widget in deleteWidget.frame.winfo_children():
                widget.destroy()
            '''
            # delete this class's widgets
            deleteWidget.frame.destroy()
        # 重新填充控件
        self.fundOperationFileInstance = []
        for number in range(numbers):
            # create numbers fundoperation view
            #print(number)
            self.fundOperationFileInstance.append(fundOperationFile(self.master))
            self.fundOperationFileInstance[number].fundoperation(number + 5)


class fundOperationFile():

    def __init__(self,master):
        print("self.frame",master)
        self.frame = tk.Frame(master)
        self.frame.pack()
        self.fundYear = tk.StringVar()
        self.fundMonth = tk.StringVar()
        self.fundDay = tk.StringVar()

        self.year = tk.StringVar()
        self.month = tk.StringVar()
        self.day = tk.StringVar()
        self.fundCodeDetailTxt = tk.StringVar()
        self.fundCode = tk.StringVar()

    #实现text动态变化
    def fundMsgDetail(self,*args):
        print(self.fundCodeChosen.get())
        if self.fundCodeChosen.get() == "310398":
            self.fundCodeDetailTxt.set("申万菱信沪深300价值指数A类")
        elif self.fundCodeChosen.get() == "090010":
            self.fundCodeDetailTxt.set("大成中证红利指数A")
        elif self.fundCodeChosen.get() == "001594":
            self.fundCodeDetailTxt.set("天弘中证银行指数A")
        elif self.fundCodeChosen.get() == "530015":
            self.fundCodeDetailTxt.set("建信深证基本面60ETF联接A")
        elif self.fundCodeChosen.get() == "000968":
            self.fundCodeDetailTxt.set("广发中证养老产业A")
        elif self.fundCodeChosen.get() == "070023":
            self.fundCodeDetailTxt.set("嘉实深证基本面120ETF联接A")
        elif self.fundCodeChosen.get() == "501021":
            self.fundCodeDetailTxt.set("华宝香港中小(QDII-LOF)A")
        elif self.fundCodeChosen.get() == "519671":
            self.fundCodeDetailTxt.set("银河沪深300价值")
        elif self.fundCodeChosen.get() == "003318":
            self.fundCodeDetailTxt.set("景顺长城中证500低波")


    def fundMsgYear(self,*args):
        # 年份获取
        print(self.yearChosen.get())
        self.fundYear.set(self.yearChosen.get())


    def fundMsgMonth(self,*args):
        # 月份获取
        print(self.monthChosen.get())
        self.fundMonth.set(self.monthChosen.get())


    def fundMsgDay(self,*args):
        # 天数获取
        print(self.dayChosen.get())
        self.fundDay.set(self.dayChosen.get())

    def fundoperation(self, arg):
        #列数值
        columnIndex = 1
        x = arg
        #print("x = ", x)
        # 定投时间
        time = tk.Label(self.frame, text="定投时间")
        time.grid(row=x, column=0)

        # 定投时间设置
        endYear = 3001 # not include 3001
        startYear = 2018
        n = endYear - startYear
        yearList=[None] * n
        yearIndex = 0
        for i in range(startYear,endYear):
            yearList[yearIndex] = i
            if i < (endYear -1):  # end year that we don't need to add 1
                yearIndex = yearIndex + 1;

        self.yearChosen = tk.ttk.Combobox(self.frame, width=5, textvariable=self.year, state='readonly')
        self.yearChosen['values'] = yearList # 设置下拉列表的值
        self.yearChosen.grid(row=x,column=columnIndex) # 设置其在界面中出现的位置 column代表列 row 代表行
        self.yearChosen.current(1) # 设置下拉列表默认显示的值，1为yearChosen['values'] 的下标值
        self.yearChosen.bind("<<ComboboxSelected>>",self.fundMsgYear)  #绑定事件,(下拉列表框被选中时，绑定fundMsgYear()函数)

        columnIndex += 1
        self.monthChosen = tk.ttk.Combobox(self.frame, width=3, textvariable=self.month, state='readonly')
        self.monthChosen['values'] = ('1','2','3','4','5','6','7','8','9','10','11','12') # 设置下拉列表的值
        self.monthChosen.grid(row=x,column=columnIndex) # 设置其在界面中出现的位置 column代表列 row 代表行
        self.monthChosen.current(0)
        self.monthChosen.bind("<<ComboboxSelected>>",self.fundMsgMonth)


        columnIndex += 1
        self.dayChosen = tk.ttk.Combobox(self.frame, width=3, textvariable=self.day, state='readonly')
        self.dayChosen['values'] = ('1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31') # 设置下拉列表的值
        self.dayChosen.grid(row=x,column=columnIndex) # 设置其在界面中出现的位置 column代表列 row 代表行
        self.dayChosen.current(0)
        self.dayChosen.bind("<<ComboboxSelected>>",self.fundMsgDay)


        columnIndex += 1
        tk.Label(self.frame, text="基金代码").grid(row=x,column=columnIndex)

        # 一个下拉列表
        columnIndex += 1
        self.fundCodeChosen = ttk.Combobox(self.frame, width=8, textvariable=self.fundCode, state='readonly')
        self.fundCodeChosen['values'] = ("310398","090010","001594","530015","000968","070023","501021","519671","003318") # 设置下拉列表的值
        self.fundCodeChosen.grid(row=x,column=5) # 设置其在界面中出现的位置 column代表列 row 代表行
        self.fundCodeChosen.current(0) # 设置下拉列表默认显示的值，0为fundCodeChosen['values'] 的下标值
        self.fundCodeChosen.bind("<<ComboboxSelected>>",self.fundMsgDetail)  #绑定事件,(下拉列表框被选中时，绑定fundMsgDetail()函数)

        # by using textvariable to implement to change the text's content automaticlly
        # 显示选择的基金品种
        columnIndex += 1
        self.fundCodeDetailTxt.set("暂没选择基金类别")
        tk.Label(self.frame,width=25, fg="red",textvariable=self.fundCodeDetailTxt).grid(row=x,column=columnIndex)

        # 买入卖出金额
        columnIndex += 1
        money = tk.Label(self.frame, text="买入卖出金额")
        money.grid(row=x, column=columnIndex)

        columnIndex += 1
        self.moneySet = tk.Entry(self.frame,width=8)
        self.moneySet.grid(row=x, column=columnIndex)

        # 成交单价
        columnIndex += 1
        self.onePrice = tk.Label(self.frame, text="成交单价")
        self.onePrice.grid(row=x, column=columnIndex)

        columnIndex += 1
        self.onePriceSet = tk.Entry(self.frame,width=8)
        self.onePriceSet.grid(row=x, column=columnIndex)

        # 盈利收益率
        columnIndex += 1
        self.earningRate = tk.Label(self.frame, text="盈利收益率")
        self.earningRate.grid(row=x, column=columnIndex)

        columnIndex += 1
        self.earningRateSet = tk.Entry(self.frame,width=8)
        self.earningRateSet.grid(row=x, column=columnIndex)

        # 市盈率
        columnIndex += 1
        self.peRatio = tk.Label(self.frame, text="市盈率")
        self.peRatio.grid(row=x, column=columnIndex)

        columnIndex += 1
        self.peRatioSet = tk.Entry(self.frame,width=8)
        self.peRatioSet.grid(row=x, column=columnIndex)
