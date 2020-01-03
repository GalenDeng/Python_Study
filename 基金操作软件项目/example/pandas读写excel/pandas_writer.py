import pandas as pd
import numpy  as np
import openpyxl

data = openpyxl.load_workbook('/Users/admin/Desktop/partyweb/webparty/code/partygaosan5frontend/pandas_writertest.xlsx')
#print(data.get_named_ranges()) # 输出工作页索引范围
#print(data.get_sheet_names()) # 输出所有工作页的名称
# 取第一张表
sheetname = data.sheetnames  #wb.sheetnames
table = data[sheetname[0]]  #  wb[sheetname]

print(table.title) # 输出表名
nrows = table.max_row # 获得行数
ncolumns = table.max_column # 获得行数
column = 2
values = ['Kevin','2020']
for value in values:
    table.cell(nrows+1,column).value = value
    column = column + 1
data.save('/Users/admin/Desktop/partyweb/webparty/code/partygaosan5frontend/pandas_writertest.xlsx')


'''
#df = pd.DataFrame({'name':["galen"],'year':["2018"] })
df = pd.read_excel('/Users/admin/Desktop/partyweb/webparty/code/partygaosan5frontend/tkinter_writertest.xlsx',sheet_name='Sheet1',engine="openpyxl")
df.loc[2] = ["galen11","2017"]
with pd.ExcelWriter('/Users/admin/Desktop/partyweb/webparty/code/partygaosan5frontend/tkinter_writertest.xlsx',mode='a',engine="openpyxl") as writer:
	df.to_excel(writer,sheet_name='Sheet1',startrow=len(df.index))
#df.to_excel('/Users/admin/Desktop/partyweb/webparty/code/partygaosan5frontend/tkinter_writertest.xlsx',startrow=len(df.index))
'''
'''
index = pd.date_range('1/1/2000', periods=8)
store = pd.HDFStore('/Users/admin/Desktop/partyweb/webparty/code/partygaosan5frontend/store.h5')
df = pd.DataFrame(np.random.randn(8, 3), index=index,columns=['A', 'B', 'C'])
store['df'] = df
'''