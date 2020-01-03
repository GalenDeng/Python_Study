import pandas as pd
import numpy  as np
import openpyxl
data = openpyxl.load_workbook('/Users/admin/Desktop/partyweb/webparty/code/partygaosan5frontend/pandas_writertest.xlsx') # 读取xlsx文件
sheetname = data.sheetnames  #wb.sheetnames
table = data[sheetname[0]]  #  wb[sheetname]
nrows = table.rows # 获得行数 类型为迭代器
ncols = table.columns # 获得列数 类型为迭代器
print(type(nrows))
for row in nrows:
    print(row) # 包含了页名，cell，值
    line = [col.value for col in row] # 取值
    print(line)
# 读取单元格
print(table.cell(1,1).value)
